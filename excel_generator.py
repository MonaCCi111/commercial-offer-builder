import json
import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage

class OfferGenerator:
    def __init__(self, template_path=None):
        self.template_path = template_path

        # Задаем цвета для подсветки
        # Красный (FFC7CE) - для флага сомнения (r == True)
        self.color_doubt = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        # Желтый (FFF2CC) - для позиций без цены (p == 0)
        self.color_no_price = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        # Корпоративный цвет шапки таблицы - F45055
        self.color_header = PatternFill(start_color="FFF45055", end_color="FFF45055", fill_type="solid")
        
        self.border_thin = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

    def generate(self, data_json_str: str, output_path="commercial_offer.xlsx") -> str:
        try:
            data = json.loads(data_json_str)
        except json.JSONDecodeError:
            raise ValueError(f"Ошибка парсинга. ИИ вернул невалидный JSON: {data_json_str}")

        if self.template_path and os.path.exists(self.template_path):
            wb = load_workbook(self.template_path)
            ws = wb.active
            start_row = 15
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "КП"
            
            # --- ШАПКА ДОКУМЕНТА ---
            # Вставка логотипа
            if getattr(sys, 'frozen', False):
                # Если программа запущена как .exe, берем путь до экзешника
                base_dir = os.path.dirname(sys.executable)
            else:
                # Если как обычный скрипт
                base_dir = os.path.dirname(__file__)

            logo_path = os.path.join(base_dir, "logo.png")

            if os.path.exists(logo_path):
                try:
                    img = OpenpyxlImage(logo_path)
                    img.width = int(img.width * (120 / max(img.height, 1)))
                    img.height = 120
                    ws.add_image(img, "A1")
                except Exception as e:
                    print(f"Не удалось добавить логотип: {e}")
            
            # Реквизиты компании
            ws.merge_cells("C1:E1")
            cell_company = ws.cell(row=1, column=3, value="ИП Дягилев И.И.")
            cell_company.font = Font(name="Arial", size=18, bold=True, color="FFF45055")
            cell_company.alignment = Alignment(horizontal="right", vertical="center")
            
            ws.merge_cells("C2:E2")
            cell_address = ws.cell(row=2, column=3, value="г. Новосибирск, ул. Дачная 60А, офис 218")
            cell_address.font = Font(name="Arial", size=11)
            cell_address.alignment = Alignment(horizontal="right", vertical="center")

            ws.merge_cells("C3:E3")
            cell_contacts = ws.cell(row=3, column=3, value="Тел: +7 (383) 209-05-40\nEmail: teplomir2090540@gmail.com")
            cell_contacts.font = Font(name="Arial", size=11, bold=True)
            cell_contacts.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

            # Заголовок документа
            current_date = datetime.now().strftime("%d.%m.%Y")
            ws.merge_cells("A6:E6")
            cell_title = ws.cell(row=6, column=1, value=f"КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ от {current_date} г.")
            cell_title.font = Font(name="Arial", size=16, bold=True)
            cell_title.alignment = Alignment(horizontal="center", vertical="center")

            start_row = 8
            headers = ["Наименование", "Ед. изм.", "Кол-во", "Цена", "Сумма", "Примечание (ИИ)"]
            
            for col_num, header_text in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_num, value=header_text)
                cell.font = Font(bold=True, color="FFFFFFFF")
                cell.fill = self.color_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self.border_thin

            start_row += 1

        current_row = start_row
        for item in data:
            name = item.get("n", "")
            unit = item.get("u", "").lower()
            quantity = item.get("q", 0)
            price = item.get("p", 0)
            doubt = item.get("r", False)
            reason = item.get("rr", "")

            ws.cell(row=current_row, column=1, value=name)
            ws.cell(row=current_row, column=2, value=unit)
            ws.cell(row=current_row, column=3, value=quantity)

            price_cell = ws.cell(row=current_row, column=4)
            sum_cell = ws.cell(row=current_row, column=5)
            note_cell = ws.cell(row=current_row, column=6)

            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.border_thin
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            if price == 0:
                price_cell.value = ""
                sum_cell.value = ""
                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).fill = self.color_no_price
            else:
                price_cell.value = price
            
            if price != 0:
                sum_cell.value = f"=C{current_row}*D{current_row}"

            if doubt:
                note_cell.value = reason
                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).fill = self.color_doubt

            current_row += 1

        # Итоги
        total_row = current_row
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
        cell_total_label = ws.cell(row=total_row, column=1, value="ИТОГО:")
        cell_total_label.font = Font(bold=True)
        cell_total_label.alignment = Alignment(horizontal="right", vertical="center")
        
        for col in range(1, 7):
            ws.cell(row=total_row, column=col).border = self.border_thin

        end_data_row = current_row - 1
        if end_data_row >= start_row:
            total_sum_formula = f"=SUM(E{start_row}:E{end_data_row})"
            cell_total_sum = ws.cell(row=total_row, column=5, value=total_sum_formula)
            cell_total_sum.font = Font(bold=True)
            cell_total_sum.number_format = '#,##0.00 ₽'

        # Подвал (Footer) удален по просьбе пользователя, оставлена только сумма в ИТОГО
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 35
        
        # Настройка высоты строк шапки (чтобы логотип влез)
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 30
        ws.row_dimensions[4].height = 20

        # Настройка параметров страницы для печати / PDF
        ws.page_setup.orientation = "portrait"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # Задаем область печати с колонки A по E (исключая колонку F с примечаниями)
        ws.print_area = f"A1:E{total_row}"

        wb.save(output_path)
        return output_path
